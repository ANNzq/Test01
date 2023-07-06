import openpyxl
import val as val

workbook = openpyxl.Workbook()
sheet1 = workbook.create_sheet("MemMap")  # 创建sheet1
sheet2 = workbook.create_sheet(".map文件中的summary")  # 创建sheet2
sheet3 = workbook.create_sheet("ROM_RAM使用率")  # 创建sheet3
sheet1.column_dimensions['A'].width = 20  # 设置A列宽度
sheet1.column_dimensions['B'].width = 20  # 设置B列宽度
sheet1.column_dimensions['C'].width = 10  # 设置C列宽度
sheet1.column_dimensions['D'].width = 10  # 设置D列宽度
sheet1.column_dimensions['E'].width = 40  # 设置E列宽度
sheet2.column_dimensions['A'].width = 20  # 设置A列宽度
sheet2.column_dimensions['B'].width = 12  # 设置B列宽度
sheet2.column_dimensions['C'].width = 12  # 设置C列宽度
sheet2.column_dimensions['D'].width = 12  # 设置D列宽度
sheet3.column_dimensions['A'].width = 10  # 设置A列宽度
fileHandler = open("E:/0602MAP文件分析工具参考/ESC_5ASGMWEQ1001AA.map", "r")
print(fileHandler)
##  文本处理
n = 1
m = 1
count = 0
row_number1 = 1
row_number2 = 1
col_number = 1

def zdw_write(x, row, sheet):
    col = 1
    for number in x:
        if number != '':
            sheet.cell(row_number1, col, number)
            col = col + 1
def zdw_write2(x, row, sheet):
    col = 1
    for number in x:
        if number != '':
            sheet.cell(row_number2, col, number)
            col = col + 1
while True:
    # Get next line from file
    line = fileHandler.readline()
    # print(line)
    # If line is empty then end of file reached
    if not line:
        break
    s = line.strip()
    print(s[0:3:1])
    if s[0:3:1] == "0x7":
        x1 = s.split(" ")
        # print(x1)
        zdw_write(x1, row_number1, sheet1)
        row_number1 = row_number1 + 1
        # sheet1.cell(n, 1).value = s
        # n = n + 1
    elif s[0:3:1] == "0x8":
        # sheet1.cell(n, 1).value = s
        # n = n + 1
        x1 = s.split(" ")
        # print(x1)
        zdw_write(x1, row_number1, sheet1)
        row_number1 = row_number1 + 1
    elif s[0:14:1] == "Start      End":
        # sheet1.cell(n, 1).value = s
        # n = n + 1
        x1 = s.split(" ")
        # print(x1)
        zdw_write(x1, row_number1, sheet1)
        row_number1 = row_number1 + 1
    if s == "----- END EXTENDED MAP LISTING -----":
        break

# 删除重复数据
sheet1_row = int((row_number1 + 1) / 2)
print(sheet1_row)
print(row_number1)
sheet1.delete_rows(sheet1_row, row_number1 - sheet1_row)

# 删除D-H列无关数据
sheet1.delete_cols(4, 5)

# 新增加一列D
sheet1.insert_cols(4)

# 将C1-E1单元格改名
sheet1.cell(1, 3).value = "RAMSize"
sheet1.cell(1, 4).value = "ROMSize"
sheet1.cell(1, 5).value = "Input object"
sheet1.cell(1, 6).value = ""

# 区分RAM和ROM
for i in range(2, sheet1_row):
    sheet1.cell(i, 3).data_type = "int"
    sheet1.cell(i, 4).data_type = "int"

k = 2
val_RAM = 0
val_ROM = 0
while True:
    if sheet1.cell(k, 2).value[0:3:1] == "0x7":
        sheet1.cell(k, 4).value = ""
        val_RAM = int(val_RAM) + int(sheet1.cell(k, 3).value)
        # print(val_RAM)
        k += 1
    elif sheet1.cell(k, 2).value[0:3:1] == "0x8":
        sheet1.cell(k, 4).value = sheet1.cell(k, 3).value
        sheet1.cell(k, 3).value = ""
        val_ROM = int(val_ROM) + int(sheet1.cell(k, 4).value)
        # print(val_ROM)
        k += 1
    if k == sheet1_row:
        break


# .map文件中的summary模块内容
while True:
    line = fileHandler.readline()
    if not line:
        break
    s = line.strip()
    x = s[-1:-11:-1]
    y = s[-1:-3:-1]
    z = s[-1:-4:-1]
    if x[-1:-11:-1] == "Attributes":
        # sheet2.cell(m, 1).value = s
        # m = m + 1
        x2 = s.split(" ")
        # print(x1)
        zdw_write2(x2, row_number2, sheet2)
        row_number2 = row_number2 + 1
    elif y[-1:-3:-1] == "!p":
        # sheet2.cell(m, 1).value = s
        # m = m + 1
        x2 = s.split(" ")
        # print(x1)
        zdw_write2(x2, row_number2, sheet2)
        row_number2 = row_number2 + 1
    elif z[-1:-4:-1] == "!xp":
        # sheet2.cell(m, 1).value = s
        # m = m + 1
        x2 = s.split(" ")
        # print(x1)
        zdw_write2(x2, row_number2, sheet2)
        row_number2 = row_number2 + 1
# 删除E-G列无关数据
sheet2.delete_cols(5, 3)
# 新增加两列，并命名
sheet2.cell(1, 5).value = "RamUsed"
sheet2.cell(1, 6).value = "RomUsed"
# 区分RAM和ROM
l = 2
while True:
    if sheet2.cell(l, 2).value[0:3:1] == "0x7":
        sheet2.cell(l, 5).value = eval(sheet2.cell(l, 4).value)
        sheet2.cell(l, 6).value = 0
        l += 1
    elif sheet2.cell(l, 2).value[0:3:1] == "0x8":
        sheet2.cell(l, 6).value = eval(sheet2.cell(l, 4).value)
        sheet2.cell(l, 5).value = 0
        l += 1
    if l == row_number2:
        break

# ROM_RAM使用率
sheet3.cell(1, 2).value = "RAM"
sheet3.cell(1, 3).value = "ROM"
sheet3.cell(2, 1).value = "Size"
sheet3.cell(3, 1).value = "Used"
sheet3.cell(4, 1).value = "Usage rate"
sheet3.cell(5, 1).value = "Total"
sheet3.cell(3, 2).value = val_RAM
sheet3.cell(3, 3).value = val_ROM

fileHandler.close()
workbook.save('E://test.xlsx')  # 保存文件
