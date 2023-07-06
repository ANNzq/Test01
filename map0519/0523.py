# import os
# from pyspark import SparkConf, SparkContext
# os.environ['PYSPARK_PYTHON'] = 'C:/Users/zhangqian/PycharmProjects/pythonProject/venv/Scripts/python.exe'
# os.environ['HADOOP_HOME'] = "D:/hadoop-3.0.0/hadoop-3.0.0"
# conf = SparkConf().setMaster("local[*]").setAppName("test_spark")
# # conf.set("spark.default.parallelism","1")
# sc = SparkContext(conf=conf)
# # 读取文件转换成RDD
# file_rdd = sc.textFile("E:/MAP_Analyze_1.txt")
# # TODO 需求1：热门size大小的TOP3
#
# result1 = file_rdd.map(lambda x: x.split("\t")).\
#     map(lambda x:x[0]).\
#     map(lambda x: (x,1)).\
#     reduceByKey(lambda a,b: a+b).\
#     sortBy(lambda x:x[1],ascending=False, numPartitions=1).\
#     take(4)
# print("需求1的结果：",result1)
#
# # import socket
# socket_client = socket.socket()
# socket_client.connect(("localhost",8888))
# socket_client.send("hello".encode("UTF-8"))
# recv_data = socket_client.recv(1024)
# print(f"服务端回复的消息是：{recv_data.decode('UTF-8')}")
# socket_client.close()
import codecs
import re

# import openpyxl as op
# import codecs
# import re
# import numpy as np
# import xlwt
# f = codecs.open('E:/MAP_Analyze_1.txt', mode='r', encoding='utf-8')  # 打开txt文件，以‘utf-8'编码读取
# line = f.readline()  # 以行的形式进行读取文件
# x = []
# count = 0
# while line:
#     count += 1
#     a = re.split(r"\s+",line) # 每行数据分隔情况，此数据以“ ”分隔
#     line_ = []
#     line_ += a
#     # line = f.readline()
#     # ff = open(r'E:/new1.xls', 'w')  # 对获取的txt前两列数据进行保存
#     wb = xlwt.Workbook()
#     ws = wb.add_sheet('sheet1')
#     # wb = xlwt.Workbook()                 # 1.创建 Workbook
#     # ws = wb.add_sheet('test_sheet')      # 2.创建 worksheet
#     ws.write(count, 0, line_[4])
#     print(count)
#     # print(line_[4])
#     ws.write(count, 1, line_[2])
#     # print(line_[2])
#     wb.save('E:/new1.xls')
# f.close()  # close文件
#
# # f = open(r'E:/new1.xls', 'w')  # 对获取的txt前两列数据进行保存
# #
# # Sheet1 = f.add_sheet('Sheet1')
# # for i in range(len(line_)):
# #     Sheet1.write(i, 0, line_[i])  # 写入数据参数对应 行, 列, 值
# # f.close()

# import codecs
# import re
# import numpy as np
# import xlwt
# f = codecs.open('E:/MAP_Analyze_1.txt', mode='r', encoding='utf-8')  # 打开txt文件，以‘utf-8'编码读取
# line = f.readline()  # 以行的形式进行读取文件
# x = []  # 设置x y z数组
# y = []
# z = []
# while line:
#     a = re.split(r"\s+",line)
#     line_ = []
#     line_ += a
# from openpyxl import Workbook
# wb = Workbook()
# ws = wb.create_sheet('hello')
# ws._current_row = 2
# # 将当前行指定在20行
# ws.append(['a','b','e'])
# # 1,2,3,4将会插入到21行，的第A,B,C,D列
# wb.save('E:/x.xlsx')
#

import pandas as pd
import openpyxl
'''
distance_list是一个列表，我们的目标是将该列表作为一列插入表格
'''
# 先打开我们的目标表格，再打开我们的目标表单
wb=openpyxl.load_workbook(r'E:/x.xlsx')
ws = wb['Sheet']
# 取出distance_list列表中的每一个元素，openpyxl的行列号是从1开始取得，所以我这里i从1开始取
f = codecs.open('E:/MAP_Analyze_1.txt', mode='r', encoding='utf-8')  # 打开txt文件，以‘utf-8'编码读取
line = f.readline()  # 以行的形式进行读取文件
x = []  # 设置x y z数组
count = 0
while line:
    a = re.split(r"\s+",line)
    count += 1
while line:
    a = re.split(r"\s+",line)
    line_ = []
    line_ += a
    for i in range(1,count):
        distance=line_[0]
        # 写入位置的行列号可以任意改变，这里我是从第2行开始按行依次插入第11列
        ws.cell(row=i+1, column=1).value =distance
    # 保存操作
    wb.save(r'E:\x.xlsx')

