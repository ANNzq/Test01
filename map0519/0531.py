import openpyxl
import xlwt
import pandas as pd
import tkinter as tk
from tkinter import filedialog
# 获取选择文件路径
# 实例化
root = tk.Tk()
root.withdraw()
# 获取文件夹路径
f_path = filedialog.askopenfilename()
print('\n获取的文件地址：', f_path)

row_number = 0
col_number = 0
# book = openpyxl.Workbook()
book = xlwt.Workbook()
count = 0
sheet = book.add_sheet('sheet1')
# sheet = book.create_sheet('sheet1')
def zdw_write(x, row, sheet):
    col = 0
    for number in x:
        if number != '':
            sheet.write(row_number, col, number)
            col = col + 1

with open(f_path) as file_object:
    for line in file_object:
        count += 1
        if "Name            " in line:
            # print(line)
            x = line.split(" ")
            # print(x)
            zdw_write(x, row_number, sheet)
            row_number = row_number + 1
            print(count)
        if "0x8" in line:
            x = line.split(" ")
            print(x)
            zdw_write(x, row_number, sheet)
            row_number = row_number + 1
print(row_number)


data = pd.read_excel('E:/book1.xls', index_col='Name')
# 查询内存大小data.loc['需要查找的Name', '内存大小Size或者Length']
data_size = data.loc['_RESET', 'Size']
print(data_size)
book.save('E:/book1.xls')